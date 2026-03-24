"""
宏观指标更新脚本
支持两种模式：
  1. 从 Excel 导入存量数据（python macro_stats_updater.py --import data.xlsx）
  2. 从国家统计局网站抓取最新月度数据，生成预览供人工校准后再写入

表：月度宏观指标
字段：指标名称 | 年月 | 数值 | 单位 | 同比变化 | 环比变化 | 数据来源 | 更新时间

指标清单（可在 INDICATORS 中扩展）：
  GDP增速、CPI同比、PPI同比、PMI制造业、PMI非制造业、
  社会消费品零售总额增速、固定资产投资增速、出口增速、进口增速、
  城镇调查失业率、工业增加值增速
"""

import os
import re
import sys
import json
import time
import logging
import argparse
from datetime import datetime, timezone, timedelta
from typing import Optional

import requests
from bs4 import BeautifulSoup
from notion_client import NotionClient, DB_IDS, SCHEMA_MACRO_STATS

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

CST = timezone(timedelta(hours=8))

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Referer": "https://data.stats.gov.cn/",
}

# ── 国家统计局数据接口配置 ────────────────────────────────────────────────────
# 使用国家统计局开放数据 API（无需鉴权的公开接口）
NBS_API = "https://data.stats.gov.cn/easyquery.htm"

INDICATORS = [
    # (指标名称, 国家统计局指标代码, 单位)
    ("GDP增速",           "A020101", "%"),
    ("CPI同比",           "A010101", "%"),
    ("PPI同比",           "A010201", "%"),
    ("PMI制造业",         "A0D0101", ""),
    ("PMI非制造业",       "A0D0201", ""),
    ("社零增速",          "A140101", "%"),
    ("固定资产投资增速",  "A140201", "%"),
    ("出口增速",          "A060101", "%"),
    ("进口增速",          "A060201", "%"),
    ("城镇调查失业率",    "A0A0101", "%"),
    ("工业增加值增速",    "A020201", "%"),
]


# ── 国家统计局抓取 ────────────────────────────────────────────────────────────

def fetch_nbs_indicator(indicator_code: str, periods: int = 13) -> list[dict]:
    """
    从国家统计局月度数据库抓取指定指标最近 N 期数据
    返回 [{"year_month": "2024-03", "value": "3.1"}, ...]
    """
    params = {
        "m": "QueryData",
        "dbcode": "hgyd",   # 月度数据库
        "rowcode": "zb",
        "colcode": "sj",
        "wds": "[]",
        "dfwds": json.dumps([{"wdcode": "zb", "valuecode": indicator_code}]),
        "k1": str(int(time.time() * 1000)),
    }
    try:
        resp = requests.get(NBS_API, params=params, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("returndata", {}).get("datanodes", [])
        results = []
        for row in rows[:periods]:
            period = row.get("wds", [{}])[0].get("valuecode", "")  # e.g. "202403"
            value = row.get("data", {}).get("strdata", "")
            if period and value:
                year_month = f"{period[:4]}-{period[4:6]}"
                results.append({"year_month": year_month, "value": value})
        return sorted(results, key=lambda x: x["year_month"])
    except Exception as e:
        logger.warning(f"国家统计局 API 失败 ({indicator_code}): {e}")
        return []


def fetch_all_latest() -> list[dict]:
    """抓取所有指标最新 13 个月数据，返回飞书记录格式"""
    records = []
    now = datetime.now(CST).strftime("%Y-%m-%d %H:%M")
    for name, code, unit in INDICATORS:
        logger.info(f"  抓取 {name} ({code})...")
        data_points = fetch_nbs_indicator(code, periods=13)
        for i, dp in enumerate(data_points):
            # 计算同比（与 12 期前对比）
            yoy = ""
            if i >= 12:
                try:
                    prev = float(data_points[i - 12]["value"])
                    curr = float(dp["value"])
                    yoy = f"{curr - prev:+.2f}"
                except Exception:
                    pass
            # 计算环比
            mom = ""
            if i >= 1:
                try:
                    prev = float(data_points[i - 1]["value"])
                    curr = float(dp["value"])
                    mom = f"{curr - prev:+.2f}"
                except Exception:
                    pass
            records.append({
                "指标名称": name,
                "年月": dp["year_month"],
                "数值": dp["value"],
                "单位": unit,
                "同比变化": yoy,
                "环比变化": mom,
                "数据来源": "国家统计局",
                "更新时间": now,
                "唯一ID": f"{name}_{dp['year_month']}",
            })
        time.sleep(0.5)
    return records


# ── Excel 导入 ────────────────────────────────────────────────────────────────

def import_from_excel(filepath: str) -> list[dict]:
    """
    从 Excel 导入存量数据
    Excel 格式要求（第一行为表头）：
      指标名称 | 年月(YYYY-MM) | 数值 | 单位 | 同比变化 | 环比变化
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [str(c.value).strip() for c in ws[1]]
    logger.info(f"Excel 表头: {headers}")

    now = datetime.now(CST).strftime("%Y-%m-%d %H:%M")
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = dict(zip(headers, row))
        name = str(row_dict.get("指标名称", "")).strip()
        year_month = str(row_dict.get("年月", "")).strip()
        value = str(row_dict.get("数值", "")).strip()
        if not name or not year_month or not value:
            continue
        records.append({
            "指标名称": name,
            "年月": year_month,
            "数值": value,
            "单位": str(row_dict.get("单位", "")).strip(),
            "同比变化": str(row_dict.get("同比变化", "")).strip(),
            "环比变化": str(row_dict.get("环比变化", "")).strip(),
            "数据来源": "Excel导入",
            "更新时间": now,
            "唯一ID": f"{name}_{year_month}",
        })
    logger.info(f"从 Excel 读取 {len(records)} 条记录")
    return records


# ── 预览模式（供人工校准）────────────────────────────────────────────────────

def preview_latest():
    """抓取最新数据并打印预览，不写入飞书，等待人工确认"""
    logger.info("=== 预览模式：抓取最新宏观指标 ===")
    records = fetch_all_latest()
    # 只显示最新一期
    latest = {}
    for r in records:
        key = r["指标名称"]
        if key not in latest or r["年月"] > latest[key]["年月"]:
            latest[key] = r

    print("\n" + "=" * 60)
    print(f"{'指标':<15} {'年月':<10} {'数值':<10} {'单位':<5} {'同比':>8} {'环比':>8}")
    print("-" * 60)
    for name, r in sorted(latest.items()):
        print(f"{name:<15} {r['年月']:<10} {r['数值']:<10} {r['单位']:<5} {r['同比变化']:>8} {r['环比变化']:>8}")
    print("=" * 60)
    print("\n✅ 请确认以上数据无误后，运行：python macro_stats_updater.py --write\n")
    return records


# ── 主入口 ────────────────────────────────────────────────────────────────────

def run_auto():
    """GitHub Actions 自动模式：抓取 + 直接写入 Notion（月度触发）"""
    logger.info("自动模式：抓取国家统计局最新数据...")
    records = fetch_all_latest()
    if not records:
        logger.warning("未获取到数据")
        return
    client = NotionClient()
    db_id = DB_IDS["macro_stats"]
    result = client.upsert_by_key(db_id, "唯一ID", records, SCHEMA_MACRO_STATS)
    logger.info(f"完成：新增 {result['created']} 条，更新 {result['updated']} 条")


def run_import(filepath: str):
    """手动导入 Excel 存量数据"""
    records = import_from_excel(filepath)
    if not records:
        logger.warning("Excel 无有效数据")
        return
    client = NotionClient()
    db_id = DB_IDS["macro_stats"]
    result = client.upsert_by_key(db_id, "唯一ID", records, SCHEMA_MACRO_STATS)
    logger.info(f"导入完成：新增 {result['created']} 条，更新 {result['updated']} 条")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="宏观指标更新工具")
    parser.add_argument("--import", dest="import_file", metavar="FILE", help="从 Excel 导入存量数据")
    parser.add_argument("--preview", action="store_true", help="预览最新数据（不写入）")
    parser.add_argument("--write", action="store_true", help="抓取并写入 Notion")
    args = parser.parse_args()

    if args.import_file:
        run_import(args.import_file)
    elif args.preview:
        preview_latest()
    elif args.write:
        run_auto()
    else:
        # GitHub Actions 默认调用
        run_auto()
